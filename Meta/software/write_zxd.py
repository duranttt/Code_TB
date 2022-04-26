#!/usr/bin/python3
import sys
import pandas as pd
import openpyxl
from openpyxl import load_workbook
table = openpyxl.load_workbook(sys.argv[1])
wb = load_workbook(filename=sys.argv[1])
sheet = table["NGS信息单"]
list_title = []
print(sheet[1][1].value)
for hang in sheet[1]:
    if hang.value is not None:
        list_title.append(hang.value)
list_number = []
for col in sheet["F"]:
    list_number.append(col.value)
position = list_title.index("宏微编号")
dic_input = {}
file2 = open(sys.argv[2],"r")
for line in file2:
    ex_name = line.strip("\n")
    table_in = pd.read_excel(ex_name)
    table_in_col = list(table_in.columns.values)
    for index, row in table_in.iterrows():
        dic_in = {}
        n = 0
        while n <= len(table_in_col)-1:
            dic_in[table_in_col[n]] = row[n]
            n = n + 1
        dic_input[row[2]] = dic_in
    for number in dic_input.keys():
        if number != number:
            continue
        else:
            pos_col = list_number.index(number) + 1
            i = 0
            for value in dic_input[number].keys():
                if value in list_title:
                    pos_row = list_title.index(value)
                    sheet.cell(pos_col,pos_row+1).value = dic_input[number][value]
                else:
                    continue
table.save(sys.argv[1])
